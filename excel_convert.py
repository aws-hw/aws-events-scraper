from openpyxl import Workbook
from openpyxl.styles import Font
import pandas as pd
import glob
import os
import datetime
import pytz
import re
from datetime import timezone, timedelta
from io import BytesIO

def convert_csv_to_excel():
    """Convert the latest AwsEvents CSV to Excel with formatted columns and NZ timezone conversion."""
    
    # Find the latest CSV file in Outputs folder
    csv_files = glob.glob('Outputs/AwsEvents_*.csv')
    if not csv_files:
        print("No CSV files found in Outputs folder")
        return
    
    latest_csv = max(csv_files, key=os.path.getctime)
    print(f"Converting {latest_csv}...")
    
    # Read CSV
    df = pd.read_csv(latest_csv)
    
    # Create workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Events"
    
    # Format column names: capitalize and remove underscores
    formatted_columns = []
    for col in df.columns:
        formatted_col = col.replace('_', ' ').title()
        # Special case for "Registration Url" -> "Event Link"
        if col == 'registration_url':
            formatted_col = 'Event Link'
        formatted_columns.append(formatted_col)
    
    # Add headers
    for col_idx, col_name in enumerate(formatted_columns, 1):
        ws.cell(row=1, column=col_idx, value=col_name)
    
    # NZ timezone (handles daylight saving automatically)
    nz_tz = pytz.timezone('Pacific/Auckland')
    
    # Add data rows
    for row_idx, row in df.iterrows():
        for col_idx, original_col in enumerate(df.columns, 1):
            cell = ws.cell(row=row_idx + 2, column=col_idx)
            value = row[original_col]
            
            # Handle registration_url column - make it a hyperlink
            if original_col == 'registration_url':
                cell.value = "Register Here"
                cell.hyperlink = value
                cell.font = Font(color="0563C1", underline="single")  # Blue underlined
            
            # Handle time column - convert to 24-hour NZ time
            elif original_col == 'time':
                try:
                    # Parse the time string
                    time_str = str(value)
                    
                    # Format is like "12:00 - 16:00 GMT+13 (10:00 - 14:00 GMT+11)"
                    # Extract the timezone offset and times
                    if 'GMT' in time_str:
                        # Get the first GMT offset mentioned
                        gmt_match = re.search(r'GMT([+-]\d+)', time_str)
                        if gmt_match:
                            offset_str = gmt_match.group(1)
                            offset_hours = int(offset_str)
                            
                            # Create a timezone with this offset
                            event_tz = timezone(timedelta(hours=offset_hours))
                            
                            # Extract start and end times (before the first GMT)
                            time_part = time_str.split('GMT')[0].strip()
                            
                            # Get the event date from the 'date' column
                            date_str = row.get('date', '')
                            event_date = datetime.date.today()
                            
                            if date_str:
                                try:
                                    # Parse date string like "Tuesday 24th February 2026"
                                    parsed_date = datetime.datetime.strptime(date_str, '%A %d%b %B %Y')
                                    event_date = parsed_date.date()
                                except:
                                    try:
                                        # Try alternative format
                                        parsed_date = datetime.datetime.strptime(date_str, '%A %dth %B %Y')
                                        event_date = parsed_date.date()
                                    except:
                                        try:
                                            # Try another format
                                            parsed_date = datetime.datetime.strptime(date_str, '%A %dst %B %Y')
                                            event_date = parsed_date.date()
                                        except:
                                            try:
                                                # Try another format
                                                parsed_date = datetime.datetime.strptime(date_str, '%A %dnd %B %Y')
                                                event_date = parsed_date.date()
                                            except:
                                                pass
                            
                            if ' - ' in time_part:
                                start_time_str, end_time_str = time_part.split(' - ')
                                start_time_str = start_time_str.strip()
                                end_time_str = end_time_str.strip()
                                
                                # Parse start time with actual event date
                                start_time_obj = datetime.datetime.strptime(start_time_str, '%H:%M').time()
                                start_dt = datetime.datetime.combine(event_date, start_time_obj, tzinfo=event_tz)
                                start_nz = start_dt.astimezone(nz_tz)
                                
                                # Parse end time with actual event date
                                end_time_obj = datetime.datetime.strptime(end_time_str, '%H:%M').time()
                                end_dt = datetime.datetime.combine(event_date, end_time_obj, tzinfo=event_tz)
                                end_nz = end_dt.astimezone(nz_tz)
                                
                                # Format as 24-hour time range
                                cell.value = f"{start_nz.strftime('%H:%M')} - {end_nz.strftime('%H:%M')}"
                            else:
                                cell.value = value
                        else:
                            cell.value = value
                    else:
                        cell.value = value
                except Exception as e:
                    print(f"Error converting time '{value}': {e}")
                    cell.value = value
            
            else:
                cell.value = value
    
    # Auto-adjust column widths
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        ws.column_dimensions[column_letter].width = min(max_length + 2, 50)
    
    # Save Excel
    timestamp = datetime.datetime.now().strftime('%Y%m%d_%H%M%S')
    excel_file = f'Outputs/AwsEvents_{timestamp}.xlsx'
    wb.save(excel_file)
    print("Done! ^_^")
    print(f"Excel file saved to {excel_file}")
    

if __name__ == '__main__':
    convert_csv_to_excel()


def convert_data_to_excel_bytes(data):
    """
    Convert scraped event data directly to Excel format in memory.
    This is used by Lambda to avoid saving files to disk.
    
    Args:
        data: List of dictionaries containing event information
              Each dict has keys: event_name, date, time, location, registration_url
              Example: [
                  {'event_name': 'AWS Summit', 'date': 'Monday 1st March 2026', ...},
                  {'event_name': 'Cloud Workshop', 'date': 'Tuesday 2nd March 2026', ...}
              ]
    
    Returns:
        bytes: Excel file content as bytes (ready to upload to S3)
    
    How it works:
        1. Takes the list of event dictionaries
        2. Creates an Excel workbook in memory (not saved to disk)
        3. Formats columns (capitalize, remove underscores)
        4. Converts registration URLs to clickable "Register Here" hyperlinks
        5. Converts times to NZ timezone in 24-hour format
        6. Returns the Excel file as bytes
    """
    from io import BytesIO
    
    # Step 1: Convert list of dicts to pandas DataFrame for easier manipulation
    df = pd.DataFrame(data)
    
    # Step 2: Create workbook in memory (not on disk)
    wb = Workbook()
    ws = wb.active
    ws.title = "Events"
    
    # Step 3: Format column names
    # Before: ['event_name', 'date', 'time', 'location', 'registration_url']
    # After: ['Event Name', 'Date', 'Time', 'Location', 'Event Link']
    formatted_columns = []
    for col in df.columns:
        formatted_col = col.replace('_', ' ').title()  # 'event_name' -> 'Event Name'
        if col == 'registration_url':
            formatted_col = 'Event Link'  # Special case
        formatted_columns.append(formatted_col)
    
    # Step 4: Write headers to first row
    for col_idx, col_name in enumerate(formatted_columns, 1):
        ws.cell(row=1, column=col_idx, value=col_name)
    
    # Step 5: Set up NZ timezone for time conversion
    nz_tz = pytz.timezone('Pacific/Auckland')
    
    # Step 6: Write data rows (starting from row 2)
    for row_idx, row in df.iterrows():
        for col_idx, original_col in enumerate(df.columns, 1):
            cell = ws.cell(row=row_idx + 2, column=col_idx)
            value = row[original_col]
            
            # Special handling for registration_url: make it a clickable hyperlink
            if original_col == 'registration_url':
                cell.value = "Register Here"  # Display text
                cell.hyperlink = value  # Actual URL
                cell.font = Font(color="0563C1", underline="single")  # Blue and underlined
            
            # Special handling for time: convert to NZ timezone
            elif original_col == 'time':
                try:
                    time_str = str(value)
                    
                    # Time format from website: "12:00 - 16:00 GMT+13 (10:00 - 14:00 GMT+11)"
                    if 'GMT' in time_str:
                        # Extract GMT offset (e.g., +13)
                        gmt_match = re.search(r'GMT([+-]\d+)', time_str)
                        if gmt_match:
                            offset_str = gmt_match.group(1)
                            offset_hours = int(offset_str)
                            
                            # Create timezone with this offset
                            event_tz = timezone(timedelta(hours=offset_hours))
                            
                            # Extract time range (before GMT)
                            time_part = time_str.split('GMT')[0].strip()
                            
                            # Get event date for accurate DST conversion
                            date_str = row.get('date', '')
                            event_date = datetime.date.today()
                            
                            if date_str:
                                # Try multiple date formats
                                for fmt in ['%A %d%b %B %Y', '%A %dth %B %Y', '%A %dst %B %Y', '%A %dnd %B %Y', '%A %drd %B %Y']:
                                    try:
                                        parsed_date = datetime.datetime.strptime(date_str, fmt)
                                        event_date = parsed_date.date()
                                        break
                                    except:
                                        continue
                            
                            # Parse start and end times
                            if ' - ' in time_part:
                                start_time_str, end_time_str = time_part.split(' - ')
                                start_time_str = start_time_str.strip()
                                end_time_str = end_time_str.strip()
                                
                                # Convert start time to NZ timezone
                                start_time_obj = datetime.datetime.strptime(start_time_str, '%H:%M').time()
                                start_dt = datetime.datetime.combine(event_date, start_time_obj, tzinfo=event_tz)
                                start_nz = start_dt.astimezone(nz_tz)
                                
                                # Convert end time to NZ timezone
                                end_time_obj = datetime.datetime.strptime(end_time_str, '%H:%M').time()
                                end_dt = datetime.datetime.combine(event_date, end_time_obj, tzinfo=event_tz)
                                end_nz = end_dt.astimezone(nz_tz)
                                
                                # Format as 24-hour time
                                cell.value = f"{start_nz.strftime('%H:%M')} - {end_nz.strftime('%H:%M')}"
                            else:
                                cell.value = value
                        else:
                            cell.value = value
                    else:
                        cell.value = value
                except Exception as e:
                    print(f"Error converting time '{value}': {e}")
                    cell.value = value
            
            # All other columns: just copy the value
            else:
                cell.value = value
    
    # Step 7: Auto-adjust column widths for readability
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        ws.column_dimensions[column_letter].width = min(max_length + 2, 50)
    
    # Step 8: Save to BytesIO (in-memory buffer, not a file)
    excel_buffer = BytesIO()
    wb.save(excel_buffer)
    excel_buffer.seek(0)  # Move pointer back to start
    
    # Step 9: Return the Excel file as bytes
    return excel_buffer.getvalue()


def convert_data_to_excel_bytes(data_list):
    """
    Convert scraped data directly to Excel format in memory.
    
    Args:
        data_list: List of dictionaries with keys: event_name, date, time, location, registration_url
        
    Returns:
        BytesIO object containing the Excel file (ready to upload to S3)
    """
    
    # Convert list of dicts to pandas DataFrame
    # DataFrame is like a table - makes it easy to work with rows and columns
    df = pd.DataFrame(data_list)
    
    # Create a new Excel workbook in memory
    wb = Workbook()
    ws = wb.active
    ws.title = "Events"
    
    # Format column names: capitalize and remove underscores
    # Example: "event_name" becomes "Event Name"
    formatted_columns = []
    for col in df.columns:
        formatted_col = col.replace('_', ' ').title()
        # Special case: "registration_url" becomes "Event Link"
        if col == 'registration_url':
            formatted_col = 'Event Link'
        formatted_columns.append(formatted_col)
    
    # Write headers to first row
    for col_idx, col_name in enumerate(formatted_columns, 1):
        ws.cell(row=1, column=col_idx, value=col_name)
    
    # NZ timezone - automatically handles daylight saving
    nz_tz = pytz.timezone('Pacific/Auckland')
    
    # Write data rows (starting from row 2)
    for row_idx, row in df.iterrows():
        for col_idx, original_col in enumerate(df.columns, 1):
            cell = ws.cell(row=row_idx + 2, column=col_idx)
            value = row[original_col]
            
            # Handle registration_url column - make it a clickable hyperlink
            if original_col == 'registration_url':
                cell.value = "Register Here"
                cell.hyperlink = value
                cell.font = Font(color="0563C1", underline="single")  # Blue underlined like a link
            
            # Handle time column - convert to 24-hour NZ time
            elif original_col == 'time':
                try:
                    time_str = str(value)
                    
                    # Parse time format like "12:00 - 16:00 GMT+13"
                    if 'GMT' in time_str:
                        # Extract the GMT offset (e.g., +13)
                        gmt_match = re.search(r'GMT([+-]\d+)', time_str)
                        if gmt_match:
                            offset_str = gmt_match.group(1)
                            offset_hours = int(offset_str)
                            
                            # Create timezone with this offset
                            event_tz = timezone(timedelta(hours=offset_hours))
                            
                            # Get time part before GMT
                            time_part = time_str.split('GMT')[0].strip()
                            
                            # Get event date from the 'date' column
                            date_str = row.get('date', '')
                            event_date = datetime.date.today()
                            
                            if date_str:
                                # Try different date formats
                                for fmt in ['%A %d%b %B %Y', '%A %dth %B %Y', '%A %dst %B %Y', '%A %dnd %B %Y', '%A %drd %B %Y']:
                                    try:
                                        parsed_date = datetime.datetime.strptime(date_str, fmt)
                                        event_date = parsed_date.date()
                                        break
                                    except:
                                        continue
                            
                            # Parse start and end times
                            if ' - ' in time_part:
                                start_time_str, end_time_str = time_part.split(' - ')
                                start_time_str = start_time_str.strip()
                                end_time_str = end_time_str.strip()
                                
                                # Convert start time to NZ timezone
                                start_time_obj = datetime.datetime.strptime(start_time_str, '%H:%M').time()
                                start_dt = datetime.datetime.combine(event_date, start_time_obj, tzinfo=event_tz)
                                start_nz = start_dt.astimezone(nz_tz)
                                
                                # Convert end time to NZ timezone
                                end_time_obj = datetime.datetime.strptime(end_time_str, '%H:%M').time()
                                end_dt = datetime.datetime.combine(event_date, end_time_obj, tzinfo=event_tz)
                                end_nz = end_dt.astimezone(nz_tz)
                                
                                # Format as 24-hour time range
                                cell.value = f"{start_nz.strftime('%H:%M')} - {end_nz.strftime('%H:%M')}"
                            else:
                                cell.value = value
                        else:
                            cell.value = value
                    else:
                        cell.value = value
                except Exception as e:
                    print(f"Error converting time '{value}': {e}")
                    cell.value = value
            
            else:
                # For all other columns, just write the value as-is
                cell.value = value
    
    # Auto-adjust column widths for readability
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        # Set width (max 50 to prevent super wide columns)
        ws.column_dimensions[column_letter].width = min(max_length + 2, 50)
    
    # Save workbook to BytesIO (in-memory file)
    excel_buffer = BytesIO()
    wb.save(excel_buffer)
    excel_buffer.seek(0)  # Reset pointer to beginning so it can be read
    
    return excel_buffer
